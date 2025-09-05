from openpyxl import load_workbook
import warnings
from collections.abc import Generator



class XlsxFile:
    """
    A helper class for reading data from an XLSX file using openpyxl.

    Provides convenient methods to access:
    - column_names(): the header row of the file.
    - data_rows(): the remaining rows of data.
    """

    def __init__(self, filepath: str) -> None:
        """
        Initialize a xlsx processor with *.xlsx file.

        Args:
            filepath(str): path to XLSX file to be read.
        """
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        self._workbook = load_workbook(filepath, data_only=True)
        self._sheet = self._workbook.active



    def column_names(self) -> tuple[str, ...]:
        """
        Return the tuple of column names (the first row of the file).

        Example:
            >>> x = XlsxFile("webpage.xlsx")
            >>> x.column_names()
            ('url|cn', 'html|cn')
        """
        reader = self._reader_generator()  # generator of excel file
        column_names_row = next(reader, ()) # this row contains column names
        return tuple(cell for cell in column_names_row if cell) # remove empty column names and return the list of column names.



    def data_rows(self) -> tuple[tuple[str | None, ...], ...]:
        """
        Return all data rows from the XLSX file (excluding the header row).

        Example:
            >>> x = XlsxFile("webpage.xlsx")
            >>> rows = x.data_rows()
            >>> rows[0][:2]
            ['abc', 'cde']
        """
        data = []

        reader = self._reader_generator() # generator of excel file

        next(reader, None) # skip the row with column names
        
        for row in reader:
            data.append(row)
        
        return tuple(data)
    


    def _reader_generator(self) -> Generator[tuple, None, None]:
        """
        Return reader generator to read sheet by row.
        """
        return self._sheet.iter_rows(values_only=True)



    def close(self) -> None:
        """
        Close the underlying workbook.
        """
        self._workbook.close()



    def __enter__(self):
        return self
    

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        if self._workbook:
            self._workbook.close()
        return True
