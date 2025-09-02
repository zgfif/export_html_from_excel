from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


class Xlsx:
    """
    A helper class for reading data from an XLSX file using openpyxl.

    Provides convenient methods to access:
    - column_names(): the header row of the file.
    - data_rows(): the remaining rows of data.
    """

    def __init__(self, filepath: str) -> None:
        """
        Initialize a xlsx processor with *.xlsx file.

        Args:
            filepath(str): path to XLSX file to be read.
        """
        self._workbook = load_workbook(filepath, data_only=True)


    def column_names(self) -> list[str | None]:
        """
        Return the list of column names (the first row of the file).

        Example:
            >>> x = Xlsx("webpage.xlsx")
            >>> x.column_names()
            ['url|cn', 'html|cn']
        """

        # select excel file sheet
        sheet = self.workbook.active  # or wb["sheet_name"]

        reader = sheet.iter_rows(values_only=True)
        row = next(reader, None) # this row contains column names

        return [cell for cell in row if cell]  # remove empty column names and return the list of column names.


    def data_rows(self) -> list[list[str | None]]:
        """
        Return all data rows from the XLSX file (excluding the header row).

        Example:
            >>> x = Xlsx("webpage.xlsx")
            >>> rows = x.data_rows()
            >>> rows[0][:2]
            ['abc', 'cde']
        """
        data = []
        
        sheet = self.workbook.active # we have only one sheet so choose it
        
        reader = sheet.iter_rows(values_only=True) # generator of excel file
        
        next(reader, None) # skip the row with column names
        
        for row in reader:
            data.append(row)
        
        return data


    @property
    def workbook(self) -> Workbook:
        """Return the openpyxl Workbook object for the loaded file."""
        return self._workbook
