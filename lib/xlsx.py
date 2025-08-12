from openpyxl import load_workbook



class Xlsx:
    def __init__(self, filepath: str) -> None:
        """Initialize a xlsx processor with *.xlsx file."""
        self._work_book = load_workbook(filepath, data_only=True)


    def column_names(self) -> list[str]:
        """Return the list of Excel column names."""
        # select excel file sheet
        sheet = self._work_book.active  # or wb["sheet_name"]

        reader = sheet.iter_rows(values_only=True)
        row = next(reader, None) # this row contains column names

        return [cell for cell in row if cell]  # remove empty column names and return the list of column names.


    def data_rows(self) ->list[list[str]]:
        """Return all the rows from Excel file."""
        data = []
        sheet = self._work_book.active # we have only one sheet so choose it
        reader = sheet.iter_rows(values_only=True) # generator of excel file
        next(reader, None) # skip the row with column names
        for row in reader:
            data.append(row)
        
        return data